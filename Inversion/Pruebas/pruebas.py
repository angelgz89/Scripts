
import os
import os.path
import openpyxl
import subprocess
import sys

if os.name == "posix":
    os.system('clear')
if os.name == "nt":
    os.system('cls')

def buscarArchivo(archivo):
    
    enc = 0
    
    if os.name == "posix":
        dir_base = "/"
    if os.name == "nt":    
        letra = str(sys.argv[0])[0:2]  # Buscar la letra de la unidad
        dir_base = letra + "\\" # Sumarle las barras para poner como directorio base la letra de la unidad
    
    for root, folders, files in os.walk(dir_base):
        for file in files:
            if file == archivo:
                enc=1
                encontrado = os.path.join(root, file)
    if enc == 1:
        return encontrado
    else:
        print("no se ha encontrado el archivo")
        return "-1"


rutaexcel = buscarArchivo("pruebasexcel.xlsm")


Libro3 = openpyxl.load_workbook(rutaexcel, keep_vba=True)
hoja3 = Libro3['Hoja']
hoja3['A3'] = "=10+20"
Libro3.save(rutaexcel)
Libro3.close()

Libro1 = openpyxl.load_workbook(rutaexcel, data_only=True, keep_vba=True)
hoja1 = Libro1['Hoja']
dine = hoja1['A1'].value
Libro1.close()


Libro2 = openpyxl.load_workbook(rutaexcel, keep_vba=True)
hoja2 = Libro2['Hoja']
hoja2['A2'] = 10*dine
Libro2.save(rutaexcel)
Libro2.close()


Libro3 = openpyxl.load_workbook(rutaexcel, data_only=True, keep_vba=True)
hoja3 = Libro3['Hoja']
dinero = hoja3["A2"].value
Libro3.close()



if os.name == "posix":
    os.system('open ' + rutaexcel)
if os.name == "nt":
    os.system('start ' + rutaexcel)